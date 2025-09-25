# core/mougli_core.py
# Mougli core + Worker CLI robusto
from __future__ import annotations

import io
import os
import sys
import json
import csv
import gc
import argparse
import traceback
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd

# ───────────────────────────── Config persistente ─────────────────────────────
APP_DIR = Path(__file__).parent
CONFIG_PATH = APP_DIR / "factores_config.json"

_DEFAULT_MONITOR = {"TV": 0.255, "CABLE": 0.425, "RADIO": 0.425, "REVISTA": 0.14875, "DIARIOS": 0.14875}
_DEFAULT_OUTVIEW = {"tarifa_superficie_factor": 1.25}


def _load_cfg() -> dict:
    if CONFIG_PATH.exists():
        try:
            cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            cfg.setdefault("monitor", _DEFAULT_MONITOR.copy())
            cfg.setdefault("outview", _DEFAULT_OUTVIEW.copy())
            cfg["outview"].setdefault("tarifa_superficie_factor", 1.25)
            return cfg
        except Exception:
            pass
    cfg = {"monitor": _DEFAULT_MONITOR.copy(), "outview": _DEFAULT_OUTVIEW.copy()}
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    return cfg


def load_monitor_factors() -> Dict[str, float]:
    return _load_cfg().get("monitor", _DEFAULT_MONITOR.copy())


def save_monitor_factors(f: Dict[str, float]) -> None:
    cfg = _load_cfg()
    cfg["monitor"] = {k: float(v) for k, v in f.items()}
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def load_outview_factor() -> float:
    return float(_load_cfg().get("outview", _DEFAULT_OUTVIEW)["tarifa_superficie_factor"])


def save_outview_factor(v: float) -> None:
    cfg = _load_cfg()
    cfg.setdefault("outview", _DEFAULT_OUTVIEW.copy())
    cfg["outview"]["tarifa_superficie_factor"] = float(v)
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


# ───────────────────────── Utiles generales ─────────────────────────
MESES_ES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _decode_bytes(b: bytes) -> str:
    for enc in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            return b.decode(enc)
        except Exception:
            pass
    raise ValueError("No fue posible decodificar el TXT.")


def _col_letter(idx: int) -> str:
    s = ""
    n = idx
    while n >= 0:
        s = chr(n % 26 + 65) + s
        n = n // 26 - 1
    return s


# ───────────────────────── Lectores de fuentes ──────────────────────
def _read_monitor_txt(file) -> pd.DataFrame:
    """Lee TXT de Monitor detectando cabecera con '|MEDIO|' y normaliza columnas."""
    if file is None:
        return pd.DataFrame()

    raw = file.read()
    text = _decode_bytes(raw) if isinstance(raw, (bytes, bytearray)) else raw
    lines = text.splitlines()

    hdr_idx = None
    for i, l in enumerate(lines[:80]):
        if "|MEDIO|" in l.upper():
            hdr_idx = i
            break
    if hdr_idx is None:
        return pd.DataFrame({"linea": [l for l in lines if l.strip()]})

    buf = StringIO("\n".join(lines[hdr_idx:]))
    df = pd.read_csv(buf, sep="|", engine="python")

    df.columns = df.columns.astype(str).str.strip()
    df = df.dropna(axis=1, how="all")

    first_col = df.columns[0]
    if first_col.strip() in {"#", ""}:
        df = df.drop(columns=[first_col])

    cols_up = {c: c.upper() for c in df.columns}
    df.rename(columns=cols_up, inplace=True)

    if "DIA" in df.columns:
        df["DIA"] = pd.to_datetime(df["DIA"], format="%d/%m/%Y", errors="coerce").dt.normalize()
        df["AÑO"] = df["DIA"].dt.year
        df["MES"] = df["DIA"].dt.month.apply(lambda m: MESES_ES[int(m)] if pd.notnull(m) and 1 <= m <= 12 else "")
        df["SEMANA"] = df["DIA"].dt.isocalendar().week

    if "MEDIO" in df.columns:
        df["MEDIO"] = df["MEDIO"].astype(str).str.upper().str.strip()

    if "INVERSION" in df.columns:
        df["INVERSION"] = pd.to_numeric(df["INVERSION"], errors="coerce").fillna(0)

    order = [c for c in ["DIA", "AÑO", "MES", "SEMANA"] if c in df.columns]
    df = df[[*order] + [c for c in df.columns if c not in order]]
    return df


def _read_out_robusto(file) -> pd.DataFrame:
    """
    Lector robusto y liviano de OutView:
    - Carga solo columnas necesarias (reduce RAM).
    - Corrige formatos numéricos de 'Tarifa S/.' (punto de miles/coma decimal).
    - Usa openpyxl en modo read_only para XLSX grandes.
    """
    if file is None:
        return pd.DataFrame()

    REQUIRED = {
        "Fecha","Latitud","Longitud","Avenida","Nro Calle/Cuadra","Marca","Tipo Elemento",
        "Orientación de Vía","Tarifa S/.","Proveedor","Distrito","Cod.Proveedor","NombreBase",
        "Item","Versión","Categoría","Anunciante","Sector","Región","Producto","Agencia"
    }
    usecols_cb = lambda c: str(c).strip() in REQUIRED

    name = (getattr(file, "name", "") or "").lower()

    def _fix_tarifa(df: pd.DataFrame) -> pd.DataFrame:
        if "Tarifa S/." in df.columns:
            s = df["Tarifa S/."].astype(str).str.strip()
            s = s.str.replace(r"(?<=\d)\.(?=\d{3}(\D|$))", "", regex=True)  # quita punto miles
            s = s.str.replace(",", ".", regex=False)                        # coma → punto
            df["Tarifa S/."] = pd.to_numeric(s, errors="coerce")
        return df

    try:
        if name.endswith(".csv"):
            engine = None
            try:
                import pyarrow  # noqa: F401
                engine = "pyarrow"
            except Exception:
                engine = None
            try:
                file.seek(0)
            except Exception:
                pass
            df = pd.read_csv(
                file,
                usecols=usecols_cb,
                engine=engine,
                low_memory=False,
                dtype_backend="pyarrow" if engine == "pyarrow" else "numpy"
            )
            return _fix_tarifa(df)

        # XLSX read_only
        try:
            try:
                file.seek(0)
            except Exception:
                pass
            from openpyxl import load_workbook
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            ws = wb.active
            headers = None
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(x) if x is not None else "" for x in row]
                    keep_idx = [j for j, h in enumerate(headers) if h.strip() in REQUIRED]
                else:
                    if not row:
                        continue
                    rows.append([row[j] for j in keep_idx])
            try:
                wb.close()
            except Exception:
                pass
            if not rows:
                return pd.DataFrame(columns=REQUIRED)
            df = pd.DataFrame(rows, columns=[headers[j] for j in keep_idx])
            return _fix_tarifa(df)
        except Exception:
            file.seek(0)
            df = pd.read_excel(file, usecols=usecols_cb, engine="openpyxl")
            return _fix_tarifa(df)
    except Exception:
        try:
            file.seek(0)
            df = pd.read_csv(file, sep=";", usecols=usecols_cb, encoding="latin-1", low_memory=False)
            return _fix_tarifa(df)
        except Exception:
            file.seek(0)
            df = pd.read_excel(file, usecols=usecols_cb, engine="openpyxl")
            return _fix_tarifa(df)


# ───────────────────────── Transformaciones ─────────────────────────
def _aplicar_factores_monitor(df: pd.DataFrame, factores: Dict[str, float]) -> pd.DataFrame:
    if df.empty or "MEDIO" not in df.columns or "INVERSION" not in df.columns:
        return df
    df = df.copy()

    def fx(m):
        m = (str(m) or "").upper().strip()
        mapa = {
            "TV": "TV", "TELEVISION": "TV",
            "CABLE": "CABLE",
            "RADIO": "RADIO",
            "REVISTA": "REVISTA", "REVISTAS": "REVISTA",
            "DIARIO": "DIARIOS", "DIARIOS": "DIARIOS", "PRENSA": "DIARIOS",
        }
        clave = mapa.get(m, None)
        return float(factores.get(clave, 1.0)) if clave else 1.0

    df["INVERSION"] = df.apply(lambda r: r["INVERSION"] * fx(r["MEDIO"]), axis=1)
    return df


def _version_column(df: pd.DataFrame) -> str | None:
    for c in df.columns:
        if str(c).lower().startswith("vers"):
            return c
    return None


def _transform_outview_enriquecido(df: pd.DataFrame, *, factor_outview: float) -> pd.DataFrame:
    """Replica cálculos de OutView (incluye Tarifa Real $) usando claves hash para bajar RAM."""
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()

    # Fechas y derivados
    df["Fecha"] = pd.to_datetime(df.get("Fecha"), dayfirst=True, errors="coerce")
    df["AÑO"] = df["Fecha"].dt.year
    df["MES"] = df["Fecha"].dt.month.apply(lambda m: MESES_ES[int(m)] if pd.notnull(m) and 1 <= m <= 12 else "")
    df["SEMANA"] = df["Fecha"].dt.isocalendar().week
    df["_FechaDT"] = df["Fecha"]
    df["_YM"] = df["_FechaDT"].dt.to_period("M")

    # Columnas categóricas para reducir memoria
    cat_cols = [
        "Proveedor","Tipo Elemento","Distrito","Avenida","Nro Calle/Cuadra","Orientación de Vía",
        "Marca","NombreBase","Categoría","Anunciante","Sector","Región","Producto","Agencia",
        "Item","Versión"
    ]
    for c in cat_cols:
        if c in df.columns:
            df[c] = df[c].astype("category")

    ver_col = _version_column(df)

    # Helpers para obtener columnas con default vacío
    def colv(c, default=""):
        if c in df.columns:
            return df[c]
        return pd.Series([default] * len(df), index=df.index)

    # === Claves hash internas (en vez de concatenar strings gigantes) ===
    from pandas.util import hash_pandas_object

    cu_cols = ["MES","AÑO","Latitud","Longitud","Avenida","Nro Calle/Cuadra","Marca",
               "Tipo Elemento","Orientación de Vía","Tarifa S/.","Proveedor","Distrito","Cod.Proveedor"]
    cu_frame = pd.DataFrame({c: colv(c) for c in cu_cols})
    s_cu = pd.Series(hash_pandas_object(cu_frame, index=False).values, index=df.index, dtype="uint64")

    c1_cols = ["NombreBase","Proveedor","Tipo Elemento","Distrito","Orientación de Vía","Nro Calle/Cuadra",
               "Item","Versión","Latitud","Longitud","Categoría","Tarifa S/.","Anunciante","MES","AÑO","SEMANA"]
    c1_frame = pd.DataFrame({c: colv(c) for c in c1_cols})
    s_c1 = pd.Series(hash_pandas_object(c1_frame, index=False).values, index=df.index, dtype="uint64")

    # === Métricas base (map en lugar de transform para ahorrar memoria) ===
    df["Denominador"] = s_cu.map(s_cu.value_counts()).astype(int)
    if ver_col:
        nunq = df.groupby(s_cu, observed=False)[ver_col].nunique()
        df["Q versiones por elemento"] = s_cu.map(nunq).astype("Int64")

    df["+1 superficie"] = s_c1.map(s_c1.value_counts()).astype(int)

    tarifa_num = pd.to_numeric(df.get("Tarifa S/."), errors="coerce").fillna(0.0)
    first_in_piece = (df.groupby(s_c1, observed=False).cumcount() == 0)

    df["Tarifa × Superficie"] = np.where(first_in_piece, tarifa_num * df["+1 superficie"], 0.0)
    df["Tarifa × Superficie"] = (df["Tarifa × Superficie"] * float(factor_outview)) / 3.8

    # Semana en mes + Conteo mensual
    df["Semana en Mes por Código"] = (
        df.groupby([s_cu, df["_YM"]], observed=False)["_FechaDT"]
          .transform(lambda s: s.rank(method="dense").astype(int))
    )
    order_in_month = (
        df.sort_values(["_FechaDT"])
          .groupby([s_cu, df["_YM"]], observed=False).cumcount()
    )
    df["Conteo Mensual"] = (order_in_month == 0).astype(int)

    # Inversión por código único
    df_pieces = df[df["Tarifa × Superficie"] != 0]
    per_first = df_pieces.groupby(s_cu.loc[df_pieces.index], observed=False)["Tarifa × Superficie"].first()
    per_count = df_pieces.groupby(s_cu.loc[df_pieces.index], observed=False).size()
    per_value = (per_first / per_count).astype(float)
    df["Tarifa × Superficie (1ra por Código único)"] = s_cu.map(per_value)

    # === Conteos y sumas AB..AI / Z con hash (evita merges pesados) ===
    # Conteo_AB_AI
    ab_cols = ["Fecha","Proveedor","Tipo Elemento","Distrito","Avenida",
               "Nro Calle/Cuadra","Orientación de Vía","Marca"]
    ab_frame = pd.DataFrame({c: colv(c) for c in ab_cols})
    k_ab = pd.Series(hash_pandas_object(ab_frame, index=False).values, index=df.index, dtype="uint64")
    cnt_ab = k_ab.value_counts()
    df["Conteo_AB_AI"] = k_ab.map(cnt_ab).astype(int)

    # NB_EXTRAE_6_7
    if "NombreBase" in df.columns:
        df["NB_EXTRAE_6_7"] = df["NombreBase"].astype(str).str.slice(5, 12)
    else:
        df["NB_EXTRAE_6_7"] = ""

    # Conteo_Z_AB_AI  y Suma_AM_Z_AB_AI
    z_cols = ["NB_EXTRAE_6_7","Proveedor","Tipo Elemento","Distrito",
              "Avenida","Nro Calle/Cuadra","Orientación de Vía","Marca"]
    z_frame = pd.DataFrame({c: colv(c) for c in z_cols})
    k_z = pd.Series(hash_pandas_object(z_frame, index=False).values, index=df.index, dtype="uint64")
    cnt_z = k_z.value_counts()
    df["Conteo_Z_AB_AI"] = k_z.map(cnt_z).astype(int)

    df["TarifaS_div3"] = tarifa_num / 3.0
    df["TarifaS_div3_sobre_Conteo"] = df["TarifaS_div3"] / df["Conteo_AB_AI"].astype(float)

    sum_z = df.groupby(k_z, observed=False)["TarifaS_div3_sobre_Conteo"].sum()
    df["Suma_AM_Z_AB_AI"] = k_z.map(sum_z).astype(float)

    # Tope por Tipo
    tipo_to_base = {
        "BANDEROLA": 12000, "CLIP": 600, "MINIPOLAR": 1000, "PALETA": 600,
        "PANEL": 1825, "PANEL CARRETERO": 5000, "PANTALLA LED": 5400,
        "PARADERO": 800, "PRISMA": 2800, "QUIOSCO": 600, "RELOJ": 840,
        "TORRE UNIPOLAR": 3000, "TOTEM": 950, "VALLA": 600, "VALLA ALTA": 1300
    }
    tipo_up = df.get("Tipo Elemento", "").astype(str).str.upper()
    tope = tipo_up.map(tipo_to_base).astype(float) * (4.0/3.0)
    an_val = pd.to_numeric(df["Suma_AM_Z_AB_AI"], errors="coerce")
    df["TopeTipo_AQ"] = tope
    df["Suma_AM_Topada_Tipo"] = np.where(np.isnan(tope), an_val, np.minimum(an_val, tope))

    denom = pd.to_numeric(df["Conteo_Z_AB_AI"], errors="coerce")
    df["SumaTopada_div_ConteoZ"] = np.where(
        denom > 0,
        pd.to_numeric(df["Suma_AM_Topada_Tipo"], errors="coerce") / denom,
        0.0
    )
    df["Tarifa Real ($)"] = np.where(
        tipo_up == "PANTALLA LED",
        df["SumaTopada_div_ConteoZ"] * 0.4,
        df["SumaTopada_div_ConteoZ"] * 0.8
    )

    # Limpieza de columnas internas pesadas
    drop_tmp = ["Tarifa × Superficie", "_FechaDT", "_YM"]
    df.drop(columns=[c for c in drop_tmp if c in df.columns], inplace=True, errors="ignore")

    # Orden final
    base = ["Fecha", "AÑO", "MES", "SEMANA"]
    tail = [
        "Denominador",
        "Q versiones por elemento" if "Q versiones por elemento" in df.columns else None,
        "+1 superficie",
        "Tarifa × Superficie (1ra por Código único)",
        "Semana en Mes por Código","Conteo Mensual",
        "NB_EXTRAE_6_7",
        "Conteo_AB_AI","Conteo_Z_AB_AI","TarifaS_div3","TarifaS_div3_sobre_Conteo",
        "Suma_AM_Z_AB_AI","TopeTipo_AQ","Suma_AM_Topada_Tipo",
        "SumaTopada_div_ConteoZ",
        "Tarifa Real ($)"
    ]
    tail = [c for c in tail if c]
    cols = [*base] + [c for c in df.columns if c not in (*base, *tail)] + tail
    return df[cols].copy()


# ───────────────────────── Resúmenes y consolidado ─────────────────────────
_BRAND_CANDS = ["MARCA", "ANUNCIANTE", "BRAND", "CLIENTE"]
_DATE_MON = ["DIA"]
_DATE_OUT = ["Fecha", "FECHA"]


def _brands_count(df: pd.DataFrame, candidates: List[str]) -> int | None:
    for c in candidates:
        if c in df.columns:
            return int(df[c].dropna().astype(str).nunique())
    return None


def _date_range(df: pd.DataFrame, candidates: List[str]) -> Tuple[str, str] | Tuple[None, None]:
    for c in candidates:
        if c in df.columns:
            s = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
            if s.notna().any():
                return (s.min().date().isoformat(), s.max().date().isoformat())
    return (None, None)


def resumen_mougli(df: pd.DataFrame, es_monitor: bool) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame([{"Filas": 0, "Rango de fechas": "", "Marcas / Anunciantes": 0}])
    d1, d2 = _date_range(df, _DATE_MON if es_monitor else _DATE_OUT)
    r = f"{d1} - {d2}" if d1 and d2 else ""
    b = _brands_count(df, _BRAND_CANDS) or 0
    return pd.DataFrame([{"Filas": len(df), "Rango de fechas": r, "Marcas / Anunciantes": b}])


_TARGET_ORDER = [
    "FECHA","AÑO","MES","SEMANA","MEDIO","MARCA","PRODUCTO","VERSIÓN",
    "DURACIÓN","TIPO ELEMENTO","TIME / Q VERSIONES","EMISORA / DISTRITO",
    "PROGRAMA / AVENIDA","BREAK / CALLE","POS. SPOT / ORIENTACIÓN",
    "INVERSIÓN REAL","SECTOR","CATEGORÍA","ÍTEM","AGENCIA","ANUNCIANTE",
    "REGIÓN","ANCHO / LATITUD","ALTO / LONGITUD","GEN / +1 SUPERFICIE",
    "Q ELEMENTOS","EDITORA / PROVEEDOR"
]
_MONITOR_MAP = {
    "DIA":"FECHA","AÑO":"AÑO","MES":"MES","SEMANA":"SEMANA",
    "MEDIO":"MEDIO","MARCA":"MARCA","PRODUCTO":"PRODUCTO","VERSION":"VERSIÓN",
    "DURACION":"DURACIÓN","TIPO":"TIPO ELEMENTO","HORA":"TIME / Q VERSIONES",
    "EMISORA/SITE":"EMISORA / DISTRITO","PROGRAMA/TIPO DE SITE":"PROGRAMA / AVENIDA",
    "BREAK":"BREAK / CALLE","POS. SPOT":"POS. SPOT / ORIENTACIÓN",
    "INVERSION":"INVERSIÓN REAL","SECTOR":"SECTOR","CATEGORIA":"CATEGORÍA",
    "ITEM":"ÍTEM","AGENCIA":"AGENCIA","ANUNCIANTE":"ANUNCIANTE",
    "REGION/ÁMBITO":"REGIÓN","ANCHO":"ANCHO / LATITUD","ALTO":"ALTO / LONGITUD",
    "GENERO":"GEN / +1 SUPERFICIE","SPOTS":"Q ELEMENTOS","EDITORA":"EDITORA / PROVEEDOR"
}
_OUT_MAP = {
    "Fecha":"FECHA","AÑO":"AÑO","MES":"MES","SEMANA":"SEMANA","Medio":"MEDIO",
    "Marca":"MARCA","Producto":"PRODUCTO","Versión":"VERSIÓN","Duración (Seg)":"DURACIÓN",
    "Tipo Elemento":"TIPO ELEMENTO","Q versiones por elemento":"TIME / Q VERSIONES",
    "Distrito":"EMISORA / DISTRITO","Avenida":"PROGRAMA / AVENIDA",
    "Nro Calle/Cuadra":"BREAK / CALLE","Orientación de Vía":"POS. SPOT / ORIENTACIÓN",
    "Tarifa Real ($)":"INVERSIÓN REAL","Sector":"SECTOR",
    "Categoría":"CATEGORÍA","Item":"ÍTEM","Agencia":"AGENCIA","Anunciante":"ANUNCIANTE",
    "Región":"REGIÓN","Latitud":"ANCHO / LATITUD","Longitud":"ALTO / LONGITUD",
    "+1 superficie":"GEN / +1 SUPERFICIE","Conteo Mensual":"Q ELEMENTOS",
    "Proveedor":"EDITORA / PROVEEDOR"
}


def _to_unified(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=_TARGET_ORDER)
    df = df.copy()
    df.rename(columns={k: v for k, v in mapping.items() if k in df.columns}, inplace=True)
    for c in _TARGET_ORDER:
        if c not in df.columns:
            df[c] = np.nan
    df = df[_TARGET_ORDER]
    if "FECHA" in df.columns:
        df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce").dt.date
    return df


# ───────────────────────── Excel (encabezado + tabla) ─────────────────────────
def _header_rows_for(df: pd.DataFrame, *, fecha_col: str | None, marca_col: str | None,
                     extras: List[Tuple[str, str]] | None = None) -> List[Tuple[str, str]]:
    filas = [("Filas", len(df))]
    if fecha_col and fecha_col in df.columns and not df.empty:
        fmin, fmax = df[fecha_col].min(), df[fecha_col].max()
        val = (f"{pd.to_datetime(fmin, errors='coerce'):%d/%m/%Y} - "
               f"{pd.to_datetime(fmax, errors='coerce'):%d/%m/%Y}") if pd.notna(fmin) else "—"
        filas.append(("Rango de fechas", val))
    if marca_col and marca_col in df.columns:
        filas.append(("Marcas / Anunciantes", df[marca_col].dropna().nunique()))
    if extras:
        for col, tit in extras:
            if col in df.columns:
                vals = ", ".join(sorted(map(str, df[col].dropna().astype(str).unique())))
                filas.append((tit, vals if vals else "—"))
    return filas


def _write_sheet_with_header_and_table(writer: pd.ExcelWriter, *,
                                       sheet_name: str,
                                       df: pd.DataFrame,
                                       header_rows: List[Tuple[str, str]]):
    header_df = pd.DataFrame(header_rows, columns=["Descripción", "Valor"])
    header_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
    ws = writer.sheets[sheet_name]

    ws.set_default_row(15)
    ws.set_row(0, 15)
    ws.set_row(1, 15)

    start_row = len(header_df) + 2
    df = df.copy()
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

    nrow, ncol = df.shape
    start_col_letter = "A"
    end_col_letter = _col_letter(ncol - 1)
    rng = f"{start_col_letter}{start_row+1}:{end_col_letter}{start_row + max(nrow,1) + 1}"

    ws.add_table(rng, {
        "name": f"{sheet_name.replace(' ', '_')}_tbl",
        "header_row": True,
        "style": "Table Style Medium 9",
        "columns": [{"header": str(h)} for h in df.columns]
    })
    ws.freeze_panes(start_row + 1, 0)
    ws.set_column(0, max(0, ncol - 1), 18)


# ───────────────────────── Función principal (UI directa) ─────────────────────
def procesar_monitor_outview(monitor_file, out_file, factores: Dict[str, float] | None,
                             outview_factor: float | None = None):
    """
    Devuelve (df_result, xlsx_bytes) con hojas condicionales:
      - 'Monitor' solo si hay Monitor
      - 'OutView' solo si hay OutView
      - 'Consolidado' solo si hay ambas fuentes
    """
    factores = factores or load_monitor_factors()
    outview_factor = float(outview_factor if outview_factor is not None else load_outview_factor())

    # MONITOR
    df_m = _read_monitor_txt(monitor_file)
    df_m = _aplicar_factores_monitor(df_m, factores)

    # OUTVIEW (enriquecido)
    df_o_raw = _read_out_robusto(out_file)
    df_o = _transform_outview_enriquecido(df_o_raw, factor_outview=outview_factor) if not df_o_raw.empty else pd.DataFrame()

    # CONSOLIDADO unificado SOLO si existen ambas fuentes
    if not df_m.empty and not df_o.empty:
        mon_u = _to_unified(df_m, _MONITOR_MAP)
        out_u = _to_unified(df_o, _OUT_MAP)
        df_c = pd.concat([mon_u, out_u], ignore_index=True)
        df_c.sort_values(["FECHA", "MARCA"], inplace=True, na_position="last")
    else:
        df_c = pd.DataFrame()

    # OutView para Excel/vista pública: ocultar internas
    internal_targets = {
        "código único","código +1 pieza","denominador",
        "tarifa × superficie","tarifa × superficie (1ra por código único)",
        "semana en mes por código",
        "nb_extrae_6_7","fecha_ab","proveedor_ac","tipoelemento_ad","distrito_ae",
        "avenida_af","nrocallecuadra_ag","orientacionvia_ah","marca_ai",
        "conteo_ab_ai","conteo_z_ab_ai","tarifas_div3","tarifas_div3_sobre_conteo",
        "suma_am_z_ab_ai","topetipo_aq","suma_am_topada_tipo","sumatopada_div_conteoz"
    }
    def _norm(s: str) -> str: return str(s).strip().casefold()
    drop_cols = [c for c in df_o.columns if _norm(c) in internal_targets]
    df_o_public = df_o.drop(columns=drop_cols, errors="ignore")

    # Excel
    xlsx = BytesIO()
    with pd.ExcelWriter(
        xlsx,
        engine="xlsxwriter",
        datetime_format="dd/mm/yyyy",
        date_format="dd/mm/yyyy",
        engine_kwargs={"options": {"constant_memory": True, "strings_to_urls": False}}
    ) as w:
        if not df_m.empty:
            hdr_m = _header_rows_for(
                df_m, fecha_col="DIA", marca_col="MARCA",
                extras=[("SECTOR","Sectores"), ("CATEGORIA","Categorías"), ("REGION/ÁMBITO","Regiones")]
            )
            _write_sheet_with_header_and_table(w, sheet_name="Monitor", df=df_m, header_rows=hdr_m)

        if not df_o_public.empty:
            hdr_o = _header_rows_for(
                df_o_public, fecha_col="Fecha", marca_col="Anunciante",
                extras=[("Tipo Elemento","Tipo"), ("Proveedor","Proveedor"), ("Región","Regiones")]
            )
            _write_sheet_with_header_and_table(w, sheet_name="OutView", df=df_o_public, header_rows=hdr_o)

        if not df_c.empty:
            d1, d2 = _date_range(df_c, ["FECHA"])
            hdr_c = [("Filas", len(df_c)),
                     ("Rango de fechas", f"{d1} - {d2}" if d1 and d2 else ""),
                     ("Fuentes incluidas", "Monitor + OutView")]
            _write_sheet_with_header_and_table(w, sheet_name="Consolidado", df=df_c, header_rows=hdr_c)

        wb = w.book
        fmt = wb.add_format({"valign": "vcenter"})
        for sh in ("Monitor", "OutView", "Consolidado"):
            if sh in w.sheets:
                ws = w.sheets[sh]
                ws.set_column(0, 0, 22, fmt)
                ws.set_column(1, 60, 18, fmt)

    xlsx.seek(0)
    if not df_c.empty:
        df_result = df_c
    elif not df_m.empty:
        df_result = df_m
    else:
        df_result = df_o_public
    return df_result, xlsx


# ───────────────────────────── Worker CLI (background) ────────────────────────
def _json_atomic_write(path: Path, payload: dict):
    tmp = path.with_suffix(path.suffix + ".tmp")
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def _log_append(log_path: Path, msg: str):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as f:
        f.write(msg.rstrip() + "\n")


def _progress(progress_path: Path, status: str, step: int, total: int, message: str):
    _json_atomic_write(progress_path, {
        "status": status, "step": step, "total": total,
        "message": message, "pid": os.getpid()
    })


def _xlsx_to_csv_stream(xlsx_path: Path, csv_writer: csv.writer) -> None:
    """Convierte la PRIMERA hoja a CSV por streaming (sin cargar todo en memoria)."""
    try:
        from openpyxl import load_workbook
    except Exception:
        df = pd.read_excel(xlsx_path, engine="openpyxl")
        for i, row in enumerate(df.itertuples(index=False, name=None)):
            if i == 0:
                csv_writer.writerow(list(df.columns))
            csv_writer.writerow(list(row))
        return

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    first = True
    for r in ws.iter_rows(values_only=True):
        if first:
            headers = [str(x) if x is not None else "" for x in r]
            csv_writer.writerow(headers)
            first = False
        else:
            csv_writer.writerow(["" if x is None else x for x in r])
    try:
        wb.close()
    except Exception:
        pass


def _combine_monitor_txt(inputs: List[Path], out_txt: Path):
    out_txt.parent.mkdir(parents=True, exist_ok=True)
    with out_txt.open("wb") as w:
        for i, p in enumerate(inputs):
            with p.open("rb") as r:
                if i > 0:
                    w.write(b"\n")
                for chunk in iter(lambda: r.read(1024 * 1024), b""):
                    w.write(chunk)


def _combine_outview_to_csv(inputs: List[Path], out_csv: Path):
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f_out:
        writer = csv.writer(f_out)
        wrote_header = False
        for p in inputs:
            name = p.name.lower()
            if name.endswith(".csv"):
                with p.open("r", encoding="utf-8", errors="ignore", newline="") as f_in:
                    reader = csv.reader(f_in)
                    for i, row in enumerate(reader):
                        if i == 0:
                            if wrote_header:
                                continue
                            wrote_header = True
                        writer.writerow(row)
            else:
                if not wrote_header:
                    pass
                _xlsx_to_csv_stream(p, writer)
                wrote_header = True


def worker_run(args):
    progress = Path(args.progress)
    logf = Path(args.log)
    outxlsx = Path(args.out_xlsx)

    try:
        total_steps = 6
        _progress(progress, "running", 0, total_steps, "Inicializando…")

        mon_inputs = [Path(x) for x in (args.monitor or []) if x]
        out_inputs = [Path(x) for x in (args.outview or []) if x]
        job_dir = Path(args.job_dir) if args.job_dir else outxlsx.parent

        comb_mon = job_dir / "combined_monitor.txt" if mon_inputs else None
        comb_out = job_dir / "combined_outview.csv" if out_inputs else None

        step = 1
        if mon_inputs:
            _progress(progress, "running", step, total_steps, "Combinando Monitor…")
            _combine_monitor_txt(mon_inputs, comb_mon)
            _log_append(logf, f"Monitor combinado en {comb_mon}")
        step += 1
        if out_inputs:
            _progress(progress, "running", step, total_steps, "Combinando OutView…")
            _combine_outview_to_csv(out_inputs, comb_out)
            _log_append(logf, f"OutView combinado en {comb_out}")
        step += 1

        _progress(progress, "running", step, total_steps, "Procesando cálculos…")
        factores = json.loads(args.factores_json) if args.factores_json else load_monitor_factors()
        out_factor = float(args.outview_factor) if args.outview_factor else load_outview_factor()

        mon_fh = open(comb_mon, "rb") if comb_mon and comb_mon.exists() else None
        out_fh = open(comb_out, "rb") if comb_out and comb_out.exists() else None

        try:
            df_result, xbytes = procesar_monitor_outview(mon_fh, out_fh, factores=factores, outview_factor=out_factor)
        finally:
            for fh in (mon_fh, out_fh):
                try:
                    if fh: fh.close()
                except Exception:
                    pass

        step += 1
        _progress(progress, "running", step, total_steps, "Generando Excel…")
        outxlsx.parent.mkdir(parents=True, exist_ok=True)
        with outxlsx.open("wb") as w:
            w.write(xbytes.getvalue())

        step += 1
        _progress(progress, "done", step, total_steps, "Completado")
        _log_append(logf, f"Éxito: {outxlsx}")
        gc.collect()

    except Exception as e:
        _log_append(logf, "ERROR:\n" + "".join(traceback.format_exception(e)))
        _progress(progress, "error", 1, 1, f"{type(e).__name__}: {e}")


# ──────────────────────────────── CLI entrypoint ──────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Mougli worker/CLI")
    parser.add_argument("--as-worker", action="store_true", help="Ejecutar en modo worker")
    parser.add_argument("--monitor", action="append", help="Ruta(s) de archivos Monitor .txt", default=[])
    parser.add_argument("--outview", action="append", help="Ruta(s) de archivos OutView .csv/.xlsx", default=[])
    parser.add_argument("--out-xlsx", required=False, help="Ruta de salida Excel")
    parser.add_argument("--progress", required=False, help="Ruta del progress.json")
    parser.add_argument("--log", required=False, help="Ruta del job.log")
    parser.add_argument("--job-dir", required=False, help="Directorio del job (para combinados)")
    parser.add_argument("--factores-json", default="", help="JSON de factores monitor")
    parser.add_argument("--outview-factor", default="", help="Factor outview")
    args = parser.parse_args()

    if args.as_worker:
        worker_run(args)
    else:
        mon = open(args.monitor[0], "rb") if args.monitor else None
        out = open(args.outview[0], "rb") if args.outview else None
        df, x = procesar_monitor_outview(mon, out, factores=None, outview_factor=None)
        print(df.head(3))
        if args.out_xlsx:
            Path(args.out_xlsx).write_bytes(x.getvalue())
